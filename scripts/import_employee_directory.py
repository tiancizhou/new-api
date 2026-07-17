#!/usr/bin/env python3
import argparse
import json
import os
from collections import Counter, defaultdict, deque
from pathlib import Path
from urllib import error, request

from openpyxl import load_workbook


def text(value):
    return "" if value is None else str(value).strip()


def integer(value):
    try:
        return int(float(text(value)))
    except ValueError:
        return 0


def normalize_department_id(value, company):
    external_id = text(value)
    company_id = text(company)
    if external_id and company_id and external_id == company_id + company_id:
        return company_id
    return external_id


def rows(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Result 1"] if "Result 1" in workbook.sheetnames else workbook.worksheets[0]
    values = sheet.iter_rows(values_only=True)
    headers = [text(value) for value in next(values)]
    for raw in values:
        row = {header: raw[index] if index < len(raw) else None for index, header in enumerate(headers)}
        if any(text(value) for value in row.values()):
            yield row


def departments(path):
    records = []
    by_id = {}
    for row in rows(path):
        company = text(row.get("hrcompid"))
        external_id = normalize_department_id(row.get("id"), company)
        if not external_id:
            continue
        record = {
            "id": external_id,
            "parent_id": normalize_department_id(row.get("parent_id"), company) or "0",
            "dept_name": text(row.get("name")) or external_id,
            "full_name": text(row.get("name")) or external_id,
            "sort": integer(row.get("sort")),
            "remark": text(row.get("remark")),
            "extract": "D" if text(row.get("is_deleted")) not in ("", "0") else "",
        }
        records.append(record)
        by_id[external_id] = record

    children = defaultdict(list)
    indegree = {record["id"]: 0 for record in records}
    for record in records:
        parent_id = record["parent_id"]
        if parent_id in indegree and parent_id != record["id"]:
            children[parent_id].append(record["id"])
            indegree[record["id"]] += 1
    queue = deque(record_id for record_id, degree in indegree.items() if degree == 0)
    ordered = []
    while queue:
        record_id = queue.popleft()
        ordered.append(by_id[record_id])
        for child_id in children[record_id]:
            indegree[child_id] -= 1
            if indegree[child_id] == 0:
                queue.append(child_id)
    ordered_ids = {record["id"] for record in ordered}
    ordered.extend(record for record in records if record["id"] not in ordered_ids)
    return ordered, set(by_id)


def employee_department(row, department_ids):
    company = text(row.get("hrcompid"))
    candidates = []
    for raw in (row.get("deptroomid"), row.get("deptid")):
        value = text(raw)
        if company and value:
            candidates.append(company + value)
        if value:
            candidates.append(value)
    for candidate in candidates:
        candidate = normalize_department_id(candidate, company)
        if candidate in department_ids:
            return candidate
    return candidates[0] if candidates else ""


def employees(path, department_ids):
    records = []
    for row in rows(path):
        external_id = text(row.get("psncode"))
        employee_no = text(row.get("ztpsncode"))
        if not external_id or not employee_no:
            continue
        records.append(
            {
                "id": external_id,
                "account": employee_no,
                "real_name": text(row.get("psnname")) or employee_no,
                "email": text(row.get("email")),
                "phone": text(row.get("telephone")) or text(row.get("workphone")),
                "dept_id": employee_department(row, department_ids),
                "post_id": text(row.get("postcode")),
                "extract": "",
            }
        )
    return records


def send(base_url, token, endpoint, payload, timeout):
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = request.Request(
        base_url.rstrip("/") + endpoint,
        data=body,
        method="POST",
        headers={"Content-Type": "application/json; charset=utf-8", "X-MDM-Token": token},
    )
    try:
        with request.urlopen(req, timeout=timeout) as response:
            result = json.loads(response.read().decode("utf-8"))
    except error.HTTPError as exc:
        raise RuntimeError(exc.read().decode("utf-8", errors="replace")) from exc
    if not result.get("success"):
        raise RuntimeError(result.get("message") or str(result))
    return result.get("data") or {}


def import_all(label, records, endpoint, args):
    actions = Counter()
    failures = []
    for index, payload in enumerate(records, 1):
        try:
            result = send(args.base_url, args.token, endpoint, payload, args.timeout)
            actions[result.get("action", "unknown")] += 1
        except Exception as exc:
            failures.append((index, payload, str(exc)))
            if not args.continue_on_error:
                raise
        if index % args.progress_every == 0:
            print(f"[{label}] {index}/{len(records)}", flush=True)
    return actions, failures


def main():
    parser = argparse.ArgumentParser(description="Import departments and employees through the dedicated MDM endpoints.")
    parser.add_argument("--dept-file", required=True, type=Path)
    parser.add_argument("--employee-file", required=True, type=Path)
    parser.add_argument("--base-url", default="http://localhost:3000")
    parser.add_argument("--token", default=os.environ.get("MDM_SYNC_TOKEN", ""))
    parser.add_argument("--timeout", type=int, default=30)
    parser.add_argument("--progress-every", type=int, default=1000)
    parser.add_argument("--continue-on-error", action="store_true")
    args = parser.parse_args()
    if not args.token:
        parser.error("--token or MDM_SYNC_TOKEN is required")

    department_records, department_ids = departments(args.dept_file)
    employee_records = employees(args.employee_file, department_ids)
    print(f"departments={len(department_records)} employees={len(employee_records)}", flush=True)

    department_actions, department_failures = import_all(
        "departments", department_records, "/getMdmInfo/deptInfo", args
    )
    employee_actions, employee_failures = import_all(
        "employees", employee_records, "/getMdmInfo/userInfo", args
    )
    print(f"department_actions={dict(department_actions)}", flush=True)
    print(f"employee_actions={dict(employee_actions)}", flush=True)
    print(f"department_failures={len(department_failures)} employee_failures={len(employee_failures)}", flush=True)
    for failure in (department_failures + employee_failures)[:20]:
        print(f"failure={failure}", flush=True)
    if department_failures or employee_failures:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
